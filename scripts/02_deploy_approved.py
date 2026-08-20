"""
02_deploy_approved.py

Script 2 of the Salesforce Field Description Automation pipeline.

Workflow:
  1. Read config.yml
  2. Find the most recent review_queue_{timestamp}.xlsx in data/
  3. Read Admin decisions from Tab A and Tab B
  4. Validate the review file — stop if any row is incomplete or invalid
  5. Deploy approved changes:
       - Experiment mode: dry-run — log decisions, write nothing to Salesforce
       - Production mode: write via Salesforce Metadata API
  6. Write write_log_{timestamp}.xlsx

See wiki/02_how_it_works.md for full pipeline documentation.
"""

import json
import logging
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── logging ───────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── paths ─────────────────────────────────────────────────────────────────────

ROOT        = Path(__file__).resolve().parent.parent
CONFIG_PATH = ROOT / "config.yml"
DATA_DIR    = ROOT / "data"

# ── constants ─────────────────────────────────────────────────────────────────

VALID_DECISIONS   = {"Approve", "Edit", "Reject"}
MAX_DESC_LENGTH   = 255    # Salesforce hard limit

# Column names expected in Tab A and Tab B of the review file
COL_OBJECT        = "Object"
COL_API_NAME      = "Field_API_Name"
COL_CURRENT_DESC  = "Current_Description"
COL_LLM_SUGGEST   = "LLM_Suggestion"
COL_DECISION      = "Decision"
COL_ADMIN_VERSION = "Admin_Version"


# ─────────────────────────────────────────────────────────────────────────────
# 1. CONFIG
# ─────────────────────────────────────────────────────────────────────────────

def load_config() -> dict:
    if not CONFIG_PATH.exists():
        log.error("config.yml not found. Copy config.example.yml to config.yml and fill in your values.")
        sys.exit(1)
    with open(CONFIG_PATH) as f:
        cfg = yaml.safe_load(f)
    log.info(f"Config loaded — mode: {cfg.get('mode', 'experiment')}")
    return cfg


# ─────────────────────────────────────────────────────────────────────────────
# 2. FIND REVIEW FILE
# ─────────────────────────────────────────────────────────────────────────────

def find_review_file() -> Path:
    """Find the most recent review_queue_{timestamp}.xlsx in data/."""
    candidates = sorted(
        DATA_DIR.glob("review_queue_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        log.error(
            "No review_queue_{timestamp}.xlsx found in data/. "
            "Run Script 1 first to generate the review file."
        )
        sys.exit(1)
    chosen = candidates[0]
    log.info(f"Review file: {chosen.name}")
    if len(candidates) > 1:
        log.warning(
            f"{len(candidates)} review files found — using the most recent. "
            f"Older files: {[p.name for p in candidates[1:]]}"
        )
    return chosen


# ─────────────────────────────────────────────────────────────────────────────
# 3. READ DECISIONS
# ─────────────────────────────────────────────────────────────────────────────

def read_decisions(review_path: Path) -> list[dict]:
    """
    Read all actionable rows from Tab A and Tab B.
    Returns list of dicts with the Admin's decision for each field.
    """
    wb = openpyxl.load_workbook(review_path, data_only=True)

    action_sheets = []
    for sheet_name in wb.sheetnames:
        if "Tab A" in sheet_name or "Tab B" in sheet_name:
            action_sheets.append(wb[sheet_name])

    if not action_sheets:
        log.error("Could not find 'Tab A' or 'Tab B' in the review file. Check the sheet names.")
        sys.exit(1)

    rows = []
    for ws in action_sheets:
        sheet_rows = _read_sheet(ws)
        log.info(f"  {ws.title}: {len(sheet_rows)} rows read")
        rows.extend(sheet_rows)

    log.info(f"Total actionable rows: {len(rows)}")
    return rows


def _read_sheet(ws) -> list[dict]:
    """Parse a worksheet into list of row dicts using header row."""
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column - 1  # 0-indexed

    required_cols = [COL_OBJECT, COL_API_NAME, COL_LLM_SUGGEST, COL_DECISION]
    for col in required_cols:
        if col not in headers:
            log.error(f"Required column '{col}' not found in sheet '{ws.title}'. Headers found: {list(headers.keys())}")
            sys.exit(1)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        def get(col_name: str) -> str:
            idx = headers.get(col_name)
            if idx is None:
                return ""
            val = row[idx] if idx < len(row) else None
            return str(val).strip() if val is not None else ""

        rows.append({
            "sheet":          ws.title,
            "object":         get(COL_OBJECT),
            "field_api_name": get(COL_API_NAME),
            "current_desc":   get(COL_CURRENT_DESC),
            "llm_suggestion": get(COL_LLM_SUGGEST),
            "decision":       get(COL_DECISION),
            "admin_version":  get(COL_ADMIN_VERSION),
        })
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# 4. VALIDATE
# ─────────────────────────────────────────────────────────────────────────────

def validate_decisions(rows: list[dict]) -> list[str]:
    """
    Validate all rows before any write begins.
    Returns list of error messages. Empty list = valid.
    """
    errors = []
    for i, row in enumerate(rows, start=2):  # row 2 onwards in Excel
        name     = row.get("field_api_name", f"<row {i}>")
        decision = row.get("decision", "")

        # Every row must have a decision
        if not decision:
            errors.append(f"Row {i} ({name}): Decision is missing.")
            continue

        # Decision must be one of the valid values
        if decision not in VALID_DECISIONS:
            errors.append(
                f"Row {i} ({name}): Invalid decision '{decision}'. "
                f"Must be one of: {', '.join(sorted(VALID_DECISIONS))}"
            )
            continue

        # Edit rows must have an Admin_Version
        if decision == "Edit":
            admin_ver = row.get("admin_version", "").strip()
            if not admin_ver:
                errors.append(f"Row {i} ({name}): Decision is 'Edit' but Admin_Version is empty.")
            elif len(admin_ver) > MAX_DESC_LENGTH:
                errors.append(
                    f"Row {i} ({name}): Admin_Version is {len(admin_ver)} characters "
                    f"(maximum is {MAX_DESC_LENGTH})."
                )

        # Approve rows — check that the LLM suggestion fits the limit
        if decision == "Approve":
            suggestion = row.get("llm_suggestion", "").strip()
            if len(suggestion) > MAX_DESC_LENGTH:
                errors.append(
                    f"Row {i} ({name}): LLM_Suggestion is {len(suggestion)} characters "
                    f"(maximum is {MAX_DESC_LENGTH}). Change decision to Edit and provide a shorter version."
                )

    return errors


# ─────────────────────────────────────────────────────────────────────────────
# 5. DEPLOY
# ─────────────────────────────────────────────────────────────────────────────

def deploy(rows: list[dict], cfg: dict) -> list[dict]:
    """
    Process decisions and write to Salesforce (or dry-run in experiment mode).
    Returns list of write log records.
    """
    mode = cfg.get("mode", "experiment")
    log_records = []

    if mode == "experiment":
        log.info("Experiment mode — dry-run: no changes will be written to Salesforce")

    for row in rows:
        decision   = row["decision"]
        name       = row["field_api_name"]
        obj        = row["object"]

        if decision == "Reject":
            log_records.append(_log_record(row, "SKIPPED", "Decision: Reject — no change written."))
            continue

        new_desc = (
            row["admin_version"].strip()
            if decision == "Edit"
            else row["llm_suggestion"].strip()
        )

        if mode == "experiment":
            log_records.append(_log_record(row, "DRY_RUN", f"Would write: {new_desc[:80]}...", new_desc))
        else:
            result = _write_to_salesforce(obj, name, new_desc, cfg)
            status = "SUCCESS" if result["ok"] else "FAILED"
            log_records.append(_log_record(row, status, result.get("message", ""), new_desc))

    success = sum(1 for r in log_records if r["write_status"] == "SUCCESS")
    dry_run = sum(1 for r in log_records if r["write_status"] == "DRY_RUN")
    skipped = sum(1 for r in log_records if r["write_status"] == "SKIPPED")
    failed  = sum(1 for r in log_records if r["write_status"] == "FAILED")

    if mode == "experiment":
        log.info(f"Dry-run complete — {dry_run} would be written, {skipped} rejected")
    else:
        log.info(f"Deployment complete — {success} written, {skipped} rejected, {failed} failed")

    return log_records


def _log_record(row: dict, status: str, message: str, new_desc: str = "") -> dict:
    return {
        "object":          row.get("object", ""),
        "field_api_name":  row.get("field_api_name", ""),
        "decision":        row.get("decision", ""),
        "old_description": row.get("current_desc", ""),
        "new_description": new_desc,
        "write_status":    status,
        "message":         message,
        "timestamp":       datetime.utcnow().isoformat(),
    }


def _write_to_salesforce(obj: str, field_api_name: str, new_desc: str, cfg: dict) -> dict:
    """
    Write a single field description to Salesforce via the Metadata API.

    TO IMPLEMENT:
      1. Install simple-salesforce: pip install simple-salesforce
      2. Authenticate using cfg['salesforce'] credentials:
             from simple_salesforce import Salesforce
             sf = Salesforce(
                 username=cfg['salesforce']['username'],
                 password=cfg['salesforce']['password'],
                 security_token=cfg['salesforce']['security_token'],
                 domain=cfg['salesforce'].get('domain', 'login'),
             )
      3. Use the Metadata API to update the field description:
             result = sf.mdapi.CustomField.update(
                 metadata={
                     'fullName': f'{obj}.{field_api_name}',
                     'description': new_desc,
                 }
             )
      4. Return {"ok": True, "message": "Written successfully"} on success
         or    {"ok": False, "message": str(error)} on failure.

    Note: The Metadata API can only update custom fields (__c). Standard fields
    are not updatable via this API. SKIPPED fields in Script 1 should already
    exclude non-writable fields before they reach this point.

    Reference: https://developer.salesforce.com/docs/atlas.en-us.api_meta.meta/api_meta/
               https://pypi.org/project/simple-salesforce/
    """
    raise NotImplementedError(
        "Salesforce Metadata API write-back is not yet implemented. "
        "See the docstring of _write_to_salesforce() for implementation guidance. "
        "To run in experiment mode (dry-run), set mode: experiment in config.yml."
    )


# ─────────────────────────────────────────────────────────────────────────────
# 6. WRITE LOG
# ─────────────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
CELL_FONT   = Font(name="Arial", size=9)
CENTER_AL   = Alignment(horizontal="center", vertical="top", wrap_text=True)
LEFT_AL     = Alignment(horizontal="left",   vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

STATUS_FILL = {
    "SUCCESS": PatternFill("solid", fgColor="D7F0D7"),
    "DRY_RUN": PatternFill("solid", fgColor="D7EEFF"),
    "SKIPPED": PatternFill("solid", fgColor="E8E8E8"),
    "FAILED":  PatternFill("solid", fgColor="FFD7D7"),
}


def write_log(log_records: list[dict]) -> Path:
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    out_path  = DATA_DIR / f"write_log_{timestamp}.xlsx"
    wb        = openpyxl.Workbook()
    ws        = wb.active
    ws.title  = "Write Log"

    headers = [
        "Object", "Field_API_Name", "Decision",
        "Old_Description", "New_Description",
        "Write_Status", "Message", "Timestamp",
    ]
    col_widths = {
        "Object": 14, "Field_API_Name": 32, "Decision": 12,
        "Old_Description": 55, "New_Description": 55,
        "Write_Status": 14, "Message": 45, "Timestamp": 22,
    }

    # Header row
    ws.row_dimensions[1].height = 28
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, CENTER_AL, THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 18)

    # Data rows
    field_map = {
        "Object":          "object",
        "Field_API_Name":  "field_api_name",
        "Decision":        "decision",
        "Old_Description": "old_description",
        "New_Description": "new_description",
        "Write_Status":    "write_status",
        "Message":         "message",
        "Timestamp":       "timestamp",
    }
    for ri, record in enumerate(log_records, 2):
        status = record.get("write_status", "")
        fill   = STATUS_FILL.get(status, PatternFill("solid", fgColor="FFFFFF"))
        for ci, h in enumerate(headers, 1):
            val = record.get(field_map[h], "")
            c   = ws.cell(row=ri, column=ci, value=val)
            c.font, c.fill, c.border = CELL_FONT, fill, THIN_BORDER
            c.alignment = CENTER_AL if h in ("Object", "Decision", "Write_Status") else LEFT_AL

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"

    # Summary tab
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 12

    from collections import Counter
    status_counts = Counter(r["write_status"] for r in log_records)
    summary_rows  = [
        ("Generated at (UTC)", timestamp),
        ("Total rows processed", len(log_records)),
        ("Written (SUCCESS)",   status_counts.get("SUCCESS", 0)),
        ("Dry-run (DRY_RUN)",   status_counts.get("DRY_RUN", 0)),
        ("Rejected (SKIPPED)",  status_counts.get("SKIPPED", 0)),
        ("Failed (FAILED)",     status_counts.get("FAILED",  0)),
    ]
    for ri, (label, val) in enumerate(summary_rows, 1):
        ws2.cell(row=ri, column=1, value=label).font = Font(name="Arial", size=9, bold=True)
        ws2.cell(row=ri, column=2, value=val).font   = Font(name="Arial", size=9)

    wb.save(out_path)
    log.info(f"Write log saved: {out_path.name}")
    return out_path


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    log.info("=" * 60)
    log.info("Script 2 — Deploy Approved")
    log.info("=" * 60)

    # 1. Config
    cfg = load_config()

    # 2. Find review file
    review_path = find_review_file()

    # 3. Read decisions
    rows = read_decisions(review_path)
    if not rows:
        log.warning("No actionable rows found in the review file. Nothing to deploy.")
        sys.exit(0)

    # 4. Validate — stop entirely if anything is wrong
    errors = validate_decisions(rows)
    if errors:
        log.error(f"Validation failed — {len(errors)} error(s) found. Nothing has been written.")
        for err in errors:
            log.error(f"  {err}")
        log.error("Fix the review file and run Script 2 again.")
        sys.exit(1)

    log.info(f"Validation passed — {len(rows)} rows ready for deployment")

    # Decision summary
    from collections import Counter
    decisions = Counter(r["decision"] for r in rows)
    log.info(
        f"Decisions: Approve={decisions.get('Approve', 0)}, "
        f"Edit={decisions.get('Edit', 0)}, "
        f"Reject={decisions.get('Reject', 0)}"
    )

    # 5. Deploy
    log_records = deploy(rows, cfg)

    # 6. Write log
    write_log(log_records)

    log.info("=" * 60)
    log.info("Script 2 complete.")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
