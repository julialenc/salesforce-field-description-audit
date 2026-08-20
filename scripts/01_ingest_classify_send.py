"""
01_ingest_classify_send.py

Script 1 of the Salesforce Field Description Automation pipeline.

Workflow:
  1. Read config.yml
  2. Ingest field metadata (experiment: JSON file | production: Tooling API)
  3. Classify every field against Rules R1-R9
  4. Save sf_classified.json
  5. Check description cache — skip fields unchanged since last run
  6. Load prompt assets
  7. Deduplicate and sort fields before batching
  8. Route to LLM: FLAGGED → Prompt A | UNCERTAIN → Prompt B | REVIEWED → Prompt C
  9. Save llm_response.json
 10. Generate review_queue_{timestamp}.xlsx
 11. Update description cache

See wiki/02_how_it_works.md for full pipeline documentation.
"""

import hashlib
import json
import logging
import os
import re
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── logging ───────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── paths ─────────────────────────────────────────────────────────────────────

ROOT          = Path(__file__).resolve().parent.parent
CONFIG_PATH   = ROOT / "config.yml"
DATA_DIR      = ROOT / "data"
PROMPTS_DIR   = ROOT / "prompts"
CACHE_PATH    = DATA_DIR / "description_cache.json"

# ── classifier constants ──────────────────────────────────────────────────────

SYSTEM_FIELD_NAMES = {
    "Id", "IsDeleted", "MasterRecordId", "CreatedDate", "CreatedById",
    "LastModifiedDate", "LastModifiedById", "SystemModstamp", "LastActivityDate",
    "LastViewedDate", "LastReferencedDate",
}

PLACEHOLDER_TOKENS = [
    "tbd", "to be defined", "to be confirmed", "to be agreed",
    "todo", "to do", "fixme", "n/a",
    "see above", "see below", "ask ", "check with",
    "deprecated", "no longer used", "no longer needed",
    "not in use", "not used", "not active", "decommissioned",
    "may no longer", "no longer be", "no longer be relevant",
    "may not be needed", "may no longer be in active use",
    "legacy - ", "legacy: ", "old field",
]

VAGUE_ANCHOR_TERMS = [
    "used for", "used by", "used to",
    "relevant", "applicable", "various", "related to",
    "for tracking", "for reporting", "for reference",
    "internal use", "internal purposes",
    "contains the ", "contains additional",
    "indicates the current",
    "stores relevant", "stores various", "stores additional",
]

# Type-mismatch patterns: (field_type_set, description_patterns)
TYPE_MISMATCH_PATTERNS = [
    (
        {"Checkbox"},
        ["enter the ", "enter a ", "type in", "enter text", "free text",
         "type the ", "write the ", "comma-separated", "list of"],
    ),
    (
        {"Number", "Currency", "Percent"},
        ["select from", "select the ", "choose from", "use the dropdown",
         "pick from", "tick ", "check this box", "check the box",
         "dropdown", "picklist"],
    ),
    (
        {"Picklist", "MultiselectPicklist"},
        ["enter the ", "enter a ", "type in", "free text", "free-text",
         "freeform", "any text"],
    ),
]

AUDIENCE_MISMATCH_PATTERNS = [
    "click ", "click here",
    "select the ", "use the dropdown",
    "use this field to", "use this to",
    "fill in ", "fill this", "fill out",
    "tick this ", "tick this box",
    "check this box", "make sure to",
    "don't forget", "leave blank if",
    "once you have", "before you ",
    "you should ", "you can ", "you need to",
    "your team", "your manager",
]

# Jargon: ALL-CAPS tokens 2-5 chars that are NOT in the allow-list
ACRONYM_ALLOW_LIST = {
    "B2B", "B2C", "EU", "UK", "US", "VAT", "ERP", "CRM", "API", "LLM",
    "AI", "ID", "HQ", "HR", "IT", "PO", "SKU", "URL", "SMS", "PDF",
    "KPI", "SLA", "ROI", "MRR", "ARR", "CSV", "JSON", "XML", "SQL",
    "ISO", "UTC", "GMT", "ETA", "ETA", "FAQ", "UI", "UX", "SaaS",
    "POS", "EDI", "KYC", "AML", "GDPR", "SEPA", "IBAN", "BIC",
    # Date format placeholders — should never trigger R7
    "YYYY", "MM", "DD", "HH", "UTC", "ISO",
}

# ─────────────────────────────────────────────────────────────────────────────
# 1. CONFIG
# ─────────────────────────────────────────────────────────────────────────────

def load_config() -> dict:
    if not CONFIG_PATH.exists():
        log.error("config.yml not found. Copy config.example.yml to config.yml and fill in your values.")
        sys.exit(1)
    with open(CONFIG_PATH) as f:
        cfg = yaml.safe_load(f)
    required = ["mode", "llm"]
    for key in required:
        if key not in cfg:
            log.error(f"config.yml is missing required key: '{key}'")
            sys.exit(1)
    log.info(f"Config loaded — mode: {cfg['mode']}")
    return cfg


# ─────────────────────────────────────────────────────────────────────────────
# 2. INGEST
# ─────────────────────────────────────────────────────────────────────────────

def ingest_fields(cfg: dict) -> list[dict]:
    mode = cfg.get("mode", "experiment")
    if mode == "experiment":
        return _ingest_from_file(cfg)
    else:
        return _ingest_from_tooling_api(cfg)


def _ingest_from_file(cfg: dict) -> list[dict]:
    path = Path(cfg.get("data_source", {}).get(
        "experiment_file", "data/sf_metadata_raw_training.json"
    ))
    if not path.is_absolute():
        path = ROOT / path
    if not path.exists():
        log.error(f"Experiment file not found: {path}")
        sys.exit(1)
    with open(path) as f:
        data = json.load(f)
    fields = data.get("fields", data) if isinstance(data, dict) else data
    log.info(f"Ingested {len(fields)} fields from {path.name}")
    return fields


def _ingest_from_tooling_api(cfg: dict) -> list[dict]:
    """
    Production ingestion via Salesforce Tooling API.

    TO IMPLEMENT:
      1. Install simple-salesforce: pip install simple-salesforce
      2. Authenticate using cfg['salesforce'] credentials
      3. For each object in cfg['objects'], run:
           SELECT QualifiedApiName, Label, DataType, Description, ...
           FROM FieldDefinition
           WHERE EntityDefinition.QualifiedApiName = '<ObjectName>'
      4. Map response to the internal field dict format:
           {
             "object":          <EntityDefinition.QualifiedApiName>,
             "field_api_name":  <QualifiedApiName>,
             "field_type":      <DataType>,
             "description":     <Description> or "",
             "picklist_values": [v["value"] for v in <PicklistValues>],
             "formula":         <CalculatedFormula> or None,
             "related_object":  <ReferenceTo[0]> or None,
           }
      5. Return the combined list across all objects.

    Reference: https://developer.salesforce.com/docs/atlas.en-us.api_tooling.meta/api_tooling/
    """
    raise NotImplementedError(
        "Tooling API ingestion is not yet implemented. "
        "See the docstring of _ingest_from_tooling_api() for implementation guidance. "
        "To run in experiment mode, set mode: experiment in config.yml."
    )


# ─────────────────────────────────────────────────────────────────────────────
# 3. CLASSIFY
# ─────────────────────────────────────────────────────────────────────────────

def classify_fields(fields: list[dict], cfg: dict) -> list[dict]:
    include_standard = cfg.get("options", {}).get("include_standard_fields", False)
    classified = []
    counts = defaultdict(int)

    for f in fields:
        result = _classify_one(f, include_standard)
        f["classifier_status"] = result["status"]
        f["rule_triggered"]    = result["rule"]
        classified.append(f)
        counts[result["status"]] += 1

    log.info(
        f"Classification complete — "
        f"FLAGGED: {counts['FLAGGED']}  "
        f"UNCERTAIN: {counts['UNCERTAIN']}  "
        f"REVIEWED: {counts['REVIEWED']}  "
        f"SKIPPED: {counts['SKIPPED']}"
    )
    return classified


def _classify_one(f: dict, include_standard: bool) -> dict:
    api_name   = f.get("field_api_name", "")
    field_type = f.get("field_type", "")
    desc       = f.get("description") or ""
    picklist   = f.get("picklist_values") or []

    # ── SKIPPED ───────────────────────────────────────────────────────────────
    if api_name in SYSTEM_FIELD_NAMES:
        return {"status": "SKIPPED", "rule": None}
    if not include_standard and not api_name.endswith("__c") and field_type in ("Id", "DateTime"):
        return {"status": "SKIPPED", "rule": None}

    # ── STEP 1: R1–R5 → FLAGGED ──────────────────────────────────────────────

    # R1 — NULL or blank
    if desc.strip() == "":
        return {"status": "FLAGGED", "rule": "R1"}

    # R2 — Echo / too short
    if _is_echo_or_too_short(desc, api_name):
        return {"status": "FLAGGED", "rule": "R2"}

    # R3 — Wrong type hint
    if _has_type_mismatch(desc, field_type):
        return {"status": "FLAGGED", "rule": "R3"}

    # R4 — Undefined picklist
    if field_type in ("Picklist", "MultiselectPicklist") and picklist:
        if not _picklist_values_mentioned(desc, picklist):
            return {"status": "FLAGGED", "rule": "R4"}

    # R5 — Stale / placeholder
    if _has_placeholder(desc):
        return {"status": "FLAGGED", "rule": "R5"}

    # ── STEP 2: R6–R9 → UNCERTAIN ────────────────────────────────────────────

    # R6 — Too long (>200 chars)
    if len(desc.strip()) > 200:
        return {"status": "UNCERTAIN", "rule": "R6"}

    # R7 — Jargon without context
    if _has_unexplained_acronym(desc):
        return {"status": "UNCERTAIN", "rule": "R7"}

    # R8 — Contradictory config
    if _has_config_contradiction(desc, field_type):
        return {"status": "UNCERTAIN", "rule": "R8"}

    # R9 — Audience mismatch
    if _has_audience_mismatch(desc):
        return {"status": "UNCERTAIN", "rule": "R9"}

    # ── STEP 3: No rule fired → REVIEWED ─────────────────────────────────────
    return {"status": "REVIEWED", "rule": None}


# ── Rule implementations ──────────────────────────────────────────────────────

def _is_echo_or_too_short(desc: str, api_name: str) -> bool:
    clean = desc.strip()
    if len(clean) < 30:
        # Don't fire R2 if the text is a placeholder — let R5 handle it
        if _has_placeholder(clean):
            return False
        return True
    # Normalise field name: strip __c, underscores, lowercase
    normalised_name = re.sub(r"__c$", "", api_name, flags=re.IGNORECASE)
    normalised_name = normalised_name.replace("_", " ").lower().strip()
    # Normalise description: strip punctuation, filler words, lowercase
    filler = {
        "indicates", "stores", "contains", "this", "the", "a", "an",
        "field", "is", "for", "of", "its", "flag", "value",
    }
    desc_words = set(
        w for w in re.sub(r"[^a-z0-9 ]", "", clean.lower()).split()
        if w not in filler
    )
    name_words = set(normalised_name.split())
    # Echo if the description's meaningful words are all contained in the field name
    if name_words and desc_words and desc_words.issubset(name_words):
        return True
    return False


def _has_type_mismatch(desc: str, field_type: str) -> bool:
    desc_lower = desc.lower()
    for type_set, patterns in TYPE_MISMATCH_PATTERNS:
        if field_type in type_set:
            for pattern in patterns:
                if pattern in desc_lower:
                    return True
    return False


def _picklist_values_mentioned(desc: str, picklist: list) -> bool:
    desc_lower = desc.lower()
    return any(v.lower() in desc_lower for v in picklist)


def _has_placeholder(desc: str) -> bool:
    desc_lower = desc.lower().strip()
    for token in PLACEHOLDER_TOKENS:
        if token in desc_lower:
            return True
    return False


def _has_unexplained_acronym(desc: str) -> bool:
    # Find all-caps tokens of 2-5 characters
    tokens = re.findall(r"\b[A-Z]{2,5}\b", desc)
    for token in tokens:
        if token in ACRONYM_ALLOW_LIST:
            continue
        # Check if immediately followed by parenthetical expansion
        pattern = re.escape(token) + r"\s*\("
        if re.search(pattern, desc):
            continue
        return True
    return False


def _has_config_contradiction(desc: str, field_type: str) -> bool:
    desc_lower = desc.lower()

    if field_type in ("Date", "DateTime"):
        patterns = ["date range", "from … to", "yyyy-mm-dd to yyyy",
                    "start and end date", "range of date", "list of date",
                    "multiple date", "start date to end date"]
        return any(p in desc_lower for p in patterns)

    if field_type == "Email":
        # Single Email field described as holding multiple addresses
        patterns = ["comma-separated", "list of email", "multiple address",
                    "multiple email", "several address", "all email",
                    "all addresses", "list of addresses", "semicolon-separated"]
        return any(p in desc_lower for p in patterns)

    if field_type == "Checkbox":
        # Checkbox described as storing text content, lists, or structured data
        patterns = ["stores a list", "records a list", "contains a list",
                    "stores the list", "stores all", "records all",
                    "stores the full", "stores the complete", "stores the history",
                    "records the history", "contains the full",
                    "stores notes", "stores feedback",
                    "written feedback", "written notes", "written comment",
                    "written text", "written summary", "written record",
                    "free-text", "free text"]
        return any(p in desc_lower for p in patterns)

    if field_type in ("Number", "Currency", "Percent"):
        patterns = ["hyphenated range", "from-to", "x to y", "band of",
                    "stored as a range", "expressed as a range"]
        return any(p in desc_lower for p in patterns)

    return False


def _has_audience_mismatch(desc: str) -> bool:
    desc_lower = desc.lower()
    return any(p in desc_lower for p in AUDIENCE_MISMATCH_PATTERNS)


# ─────────────────────────────────────────────────────────────────────────────
# 4. SAVE sf_classified.json
# ─────────────────────────────────────────────────────────────────────────────

def save_classified(fields: list[dict]) -> None:
    out = DATA_DIR / "sf_classified.json"
    with open(out, "w") as f:
        json.dump({"classified_at": datetime.utcnow().isoformat(),
                   "total": len(fields), "fields": fields}, f, indent=2)
    log.info(f"Saved sf_classified.json ({len(fields)} fields)")


# ─────────────────────────────────────────────────────────────────────────────
# 5. CACHE
# ─────────────────────────────────────────────────────────────────────────────

def load_cache() -> dict:
    if CACHE_PATH.exists():
        with open(CACHE_PATH) as f:
            return json.load(f)
    return {}


def save_cache(cache: dict) -> None:
    with open(CACHE_PATH, "w") as f:
        json.dump(cache, f, indent=2)
    log.info(f"Cache updated ({len(cache)} entries)")


def description_hash(desc: str) -> str:
    return hashlib.md5((desc or "").strip().encode()).hexdigest()


def apply_cache(fields: list[dict], cache: dict) -> tuple[list[dict], list[dict]]:
    """Split fields into (needs_llm, cached). Cached fields already have llm_result."""
    needs_llm, cached = [], []
    skipped_count = 0
    for f in fields:
        if f["classifier_status"] == "SKIPPED":
            cached.append(f)
            continue
        key = f"{f['object']}.{f['field_api_name']}"
        h   = description_hash(f.get("description", ""))
        if key in cache and cache[key].get("hash") == h:
            f["llm_result"] = cache[key].get("llm_result")
            f["cache_hit"]  = True
            cached.append(f)
            skipped_count += 1
        else:
            f["cache_hit"] = False
            needs_llm.append(f)
    log.info(f"Cache: {skipped_count} fields skipped (unchanged since last run), {len(needs_llm)} fields queued for LLM")
    return needs_llm, cached


# ─────────────────────────────────────────────────────────────────────────────
# 6. PROMPTS
# ─────────────────────────────────────────────────────────────────────────────

def load_prompts() -> dict:
    files = {
        "system":    "system_prompt.md",
        "examples":  "golden_examples.json",
        "prompt_a":  "prompt_a_flagged_fields.md",
        "prompt_b":  "prompt_b_uncertain_fields.md",
        "prompt_c":  "prompt_c_reviewed_fields.md",
    }
    prompts = {}
    for key, fname in files.items():
        path = PROMPTS_DIR / fname
        if not path.exists():
            log.error(f"Prompt file missing: {path}")
            sys.exit(1)
        with open(path) as f:
            prompts[key] = f.read()
    # Parse golden examples
    prompts["examples_parsed"] = json.loads(prompts["examples"])
    log.info("All prompt assets loaded")
    return prompts


# ─────────────────────────────────────────────────────────────────────────────
# 7. SORT + DEDUPLICATE
# ─────────────────────────────────────────────────────────────────────────────

def sort_and_deduplicate(fields: list[dict]) -> tuple[list[dict], dict]:
    """
    Sort by field_type for consistent batching.
    Deduplicate by description — identical descriptions share one LLM call.
    Returns (sorted_fields, desc_to_primary) where desc_to_primary maps
    a description to the first field that had it (the one that gets the LLM call).
    """
    llm_fields = [f for f in fields if f["classifier_status"] != "SKIPPED"]
    llm_fields.sort(key=lambda f: (f.get("field_type", ""), f.get("object", "")))

    seen_descs: dict[str, dict] = {}
    unique, duplicates = [], []
    for f in llm_fields:
        key = (f.get("description") or "").strip()
        if key and key in seen_descs:
            f["dedup_primary"] = seen_descs[key]["field_api_name"]
            duplicates.append(f)
        else:
            if key:
                seen_descs[key] = f
            f["dedup_primary"] = None
            unique.append(f)

    if duplicates:
        log.info(f"Deduplication: {len(duplicates)} fields share descriptions with earlier fields — will reuse LLM responses")

    return unique, {v["field_api_name"]: v for v in unique}


# ─────────────────────────────────────────────────────────────────────────────
# 8. BATCH + ROUTE
# ─────────────────────────────────────────────────────────────────────────────

ROUTE_MAP = {
    "FLAGGED":   "prompt_a",
    "UNCERTAIN": "prompt_b",
    "REVIEWED":  "prompt_c",
}

def build_batches(fields: list[dict], batch_size: int = 50) -> list[dict]:
    """Group by classifier_status (route), then slice into batches."""
    by_route: dict[str, list] = defaultdict(list)
    for f in fields:
        status = f["classifier_status"]
        if status in ROUTE_MAP:
            by_route[status].append(f)

    batches = []
    for status, route_key in ROUTE_MAP.items():
        group = by_route.get(status, [])
        for i in range(0, len(group), batch_size):
            batches.append({
                "status":    status,
                "route_key": route_key,
                "fields":    group[i : i + batch_size],
            })
    log.info(f"Built {len(batches)} LLM batches across {sum(len(b['fields']) for b in batches)} fields")
    return batches


def build_llm_payload(batch: dict, prompts: dict) -> list[dict]:
    """Serialize fields for the LLM — only the columns the LLM needs."""
    payload = []
    for f in batch["fields"]:
        entry = {
            "field_api_name": f.get("field_api_name"),
            "field_type":     f.get("field_type"),
            "description":    f.get("description") or "",
            "picklist_values": f.get("picklist_values") or [],
            "formula":        f.get("formula"),
            "related_object": f.get("related_object"),
        }
        if f["classifier_status"] != "REVIEWED":
            entry["rule_triggered"] = f.get("rule_triggered")
        payload.append(entry)
    return payload


# ─────────────────────────────────────────────────────────────────────────────
# 9. LLM CALLS
# ─────────────────────────────────────────────────────────────────────────────

def call_llm(payload: list[dict], batch: dict, prompts: dict, cfg: dict) -> list[dict]:
    provider = cfg["llm"].get("provider", "aoai_simulator")
    system   = _build_system_message(batch["route_key"], prompts)
    user     = _build_user_message(payload, batch["route_key"], prompts)

    if provider in ("aoai_simulator", "openai", "ollama"):
        raw = _call_openai_compatible(system, user, cfg)
    elif provider == "azure_openai":
        raw = _call_azure_openai(system, user, cfg)
    elif provider == "bedrock":
        raw = _call_bedrock(system, user, cfg)
    elif provider == "vertex_ai":
        raw = _call_vertex_ai(system, user, cfg)
    else:
        log.error(f"Unknown LLM provider: {provider}")
        sys.exit(1)

    return _parse_llm_response(raw, payload)


def _build_system_message(route_key: str, prompts: dict) -> str:
    examples_block = json.dumps(prompts["examples_parsed"], indent=2)
    return (
        prompts["system"]
        + "\n\n---\n\n## Golden Examples\n\n"
        + "The following are examples of high-quality field descriptions. "
        + "Use them to calibrate the quality bar.\n\n"
        + "```json\n" + examples_block + "\n```"
    )


def _build_user_message(payload: list[dict], route_key: str, prompts: dict) -> str:
    task = prompts[route_key]
    fields_block = json.dumps(payload, indent=2)
    return (
        task
        + "\n\n---\n\n## Fields to process\n\n"
        + "```json\n" + fields_block + "\n```"
    )


def _call_openai_compatible(system: str, user: str, cfg: dict) -> str:
    endpoint = cfg["llm"].get("endpoint", "http://localhost:8080")
    model    = cfg["llm"].get("model", "gpt-4o")
    api_key  = cfg["llm"].get("api_key", "dummy")
    temp     = cfg["llm"].get("temperature", 0.2)

    url = endpoint.rstrip("/") + "/v1/chat/completions"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {api_key}",
    }
    body = {
        "model": model,
        "temperature": temp,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user",   "content": user},
        ],
    }
    resp = requests.post(url, headers=headers, json=body, timeout=120)
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"]


def _call_azure_openai(system: str, user: str, cfg: dict) -> str:
    endpoint   = cfg["llm"].get("endpoint", "").rstrip("/")
    model      = cfg["llm"].get("model", "gpt-4o")
    api_key    = cfg["llm"].get("api_key", "")
    api_version = cfg["llm"].get("api_version", "2024-02-01")
    temp       = cfg["llm"].get("temperature", 0.2)

    url = f"{endpoint}/openai/deployments/{model}/chat/completions?api-version={api_version}"
    headers = {
        "Content-Type": "application/json",
        "api-key": api_key,
    }
    body = {
        "temperature": temp,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user",   "content": user},
        ],
    }
    resp = requests.post(url, headers=headers, json=body, timeout=120)
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"]


def _call_bedrock(system: str, user: str, cfg: dict) -> str:
    """
    Amazon Bedrock implementation.

    TO IMPLEMENT:
      1. Install boto3: pip install boto3
      2. Configure AWS credentials (environment variables or ~/.aws/credentials)
      3. Use boto3.client('bedrock-runtime') to invoke the model
      4. Map system + user messages to the Bedrock request format for your chosen model
         (Claude on Bedrock uses the Messages API; other models vary)

    Reference: https://docs.aws.amazon.com/bedrock/latest/userguide/what-is-bedrock.html
    """
    raise NotImplementedError(
        "Amazon Bedrock integration is not yet implemented. "
        "See the docstring of _call_bedrock() for implementation guidance."
    )


def _call_vertex_ai(system: str, user: str, cfg: dict) -> str:
    """
    Google Vertex AI implementation.

    TO IMPLEMENT:
      1. Install google-cloud-aiplatform: pip install google-cloud-aiplatform
      2. Authenticate via Application Default Credentials
      3. Use vertexai.generative_models.GenerativeModel to send the prompt
      4. Map system + user to Vertex AI's content format

    Reference: https://cloud.google.com/vertex-ai/generative-ai/docs/start/quickstarts
    """
    raise NotImplementedError(
        "Google Vertex AI integration is not yet implemented. "
        "See the docstring of _call_vertex_ai() for implementation guidance."
    )


def _parse_llm_response(raw: str, payload: list[dict]) -> list[dict]:
    """Parse LLM JSON response. Returns list of result dicts keyed by field_api_name."""
    # Strip markdown fences if present
    cleaned = re.sub(r"```(?:json)?", "", raw).strip().strip("`").strip()
    try:
        results = json.loads(cleaned)
    except json.JSONDecodeError as e:
        log.warning(f"LLM response JSON parse error: {e} — marking all fields in batch as llm_error")
        return [
            {"field_api_name": p["field_api_name"],
             "action": "llm_error",
             "suggested_description": p.get("description", ""),
             "reasoning": f"LLM returned unparseable JSON: {str(e)[:100]}"}
            for p in payload
        ]
    # Index by field_api_name for fast lookup
    return results if isinstance(results, list) else [results]


# ─────────────────────────────────────────────────────────────────────────────
# MAIN LLM LOOP
# ─────────────────────────────────────────────────────────────────────────────

def run_llm(fields_needing_llm: list[dict], prompts: dict, cfg: dict) -> dict:
    """
    Process all fields through the LLM.
    Returns dict: field_api_name → llm_result dict.
    """
    batch_size = cfg.get("options", {}).get("llm_batch_size", 50)
    batches    = build_batches(fields_needing_llm, batch_size)
    all_results: dict[str, dict] = {}

    for i, batch in enumerate(batches, 1):
        status = batch["status"]
        n      = len(batch["fields"])
        log.info(f"Batch {i}/{len(batches)}: {status} — {n} fields")

        payload = build_llm_payload(batch, prompts)
        try:
            results = call_llm(payload, batch, prompts, cfg)
            for r in results:
                name = r.get("field_api_name")
                if name:
                    all_results[name] = r
        except Exception as e:
            log.warning(f"Batch {i} LLM call failed: {e} — fields marked as llm_error")
            for f in batch["fields"]:
                all_results[f["field_api_name"]] = {
                    "field_api_name": f["field_api_name"],
                    "action": "llm_error",
                    "suggested_description": f.get("description", ""),
                    "reasoning": str(e)[:200],
                }
        # Brief pause between batches to avoid rate limiting
        if i < len(batches):
            time.sleep(0.5)

    log.info(f"LLM processing complete — {len(all_results)} results")
    return all_results


# ─────────────────────────────────────────────────────────────────────────────
# 10. SAVE llm_response.json
# ─────────────────────────────────────────────────────────────────────────────

def save_llm_response(all_results: dict, dedup_map: dict[str, dict]) -> None:
    """Save raw LLM output, expanding deduplicated results back to all fields."""
    # Expand dedup: any field whose dedup_primary points to another field
    # gets the same result as that field
    expanded = dict(all_results)
    for f in dedup_map.values():
        primary = f.get("dedup_primary")
        if primary and primary in expanded:
            expanded[f["field_api_name"]] = {
                **expanded[primary],
                "field_api_name": f["field_api_name"],
                "dedup_reuse": True,
            }

    out = DATA_DIR / "llm_response.json"
    with open(out, "w") as fp:
        json.dump({
            "generated_at": datetime.utcnow().isoformat(),
            "total": len(expanded),
            "results": list(expanded.values()),
        }, fp, indent=2)
    log.info(f"Saved llm_response.json ({len(expanded)} results)")
    return expanded


# ─────────────────────────────────────────────────────────────────────────────
# 11. GENERATE EXCEL REVIEW QUEUE
# ─────────────────────────────────────────────────────────────────────────────

STATUS_COLORS = {
    "FLAGGED":   "FFD7D7",
    "UNCERTAIN": "FFF3CC",
    "REVIEWED":  "D7F0D7",
    "SKIPPED":   "E8E8E8",
}
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
CELL_FONT    = Font(name="Arial", size=9)
CENTER_AL    = Alignment(horizontal="center", vertical="top", wrap_text=True)
LEFT_AL      = Alignment(horizontal="left",   vertical="top", wrap_text=True)
THIN_BORDER  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def generate_review_queue(all_fields: list[dict], llm_results: dict) -> Path:
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    out_path  = DATA_DIR / f"review_queue_{timestamp}.xlsx"
    wb        = openpyxl.Workbook()

    # Merge LLM results into fields
    for f in all_fields:
        name   = f.get("field_api_name", "")
        result = llm_results.get(name, {})
        f["llm_action"]              = result.get("action", "")
        f["llm_suggested_description"] = result.get("suggested_description", f.get("description", ""))
        f["llm_reasoning"]           = result.get("reasoning", "")

    # Tab A — FLAGGED
    tab_a = [f for f in all_fields if f["classifier_status"] == "FLAGGED"]
    # Tab B — UNCERTAIN + REVIEWED fields where LLM said "rewrite"
    tab_b = [f for f in all_fields if f["classifier_status"] == "UNCERTAIN"
             or (f["classifier_status"] == "REVIEWED" and f.get("llm_action") == "rewrite")]
    # Tab C — REVIEWED-keep + SKIPPED
    tab_c = [f for f in all_fields if f["classifier_status"] == "SKIPPED"
             or (f["classifier_status"] == "REVIEWED" and f.get("llm_action") in ("keep", "", None, "llm_error"))]

    wb.remove(wb.active)  # remove default sheet
    _write_action_tab(wb, "Tab A — FLAGGED",   tab_a, include_decision=True)
    _write_action_tab(wb, "Tab B — UNCERTAIN",  tab_b, include_decision=True)
    _write_reference_tab(wb, "Tab C — Reference", tab_c)

    _write_summary_tab(wb, tab_a, tab_b, tab_c, timestamp)

    wb.save(out_path)
    log.info(f"Review queue saved: {out_path.name}  (Tab A: {len(tab_a)}, Tab B: {len(tab_b)}, Tab C: {len(tab_c)})")
    return out_path


def _col_width(col: str) -> int:
    widths = {
        "Object": 12, "Field_API_Name": 32, "Field_Type": 16,
        "Rule": 8, "Original_Status": 14,
        "Current_Description": 55, "LLM_Suggestion": 55, "LLM_Reasoning": 40,
        "Decision": 12, "Admin_Version": 55,
    }
    return widths.get(col, 20)


def _write_action_tab(wb, title: str, fields: list[dict], include_decision: bool) -> None:
    ws = wb.create_sheet(title)
    headers = [
        "Object", "Field_API_Name", "Field_Type", "Rule", "Original_Status",
        "Current_Description", "LLM_Suggestion", "LLM_Reasoning",
        "Decision", "Admin_Version",
    ]
    _write_header_row(ws, headers)

    for ri, f in enumerate(fields, 2):
        status = f["classifier_status"]
        bg     = STATUS_COLORS.get(status, "FFFFFF")
        fill   = PatternFill("solid", fgColor=bg)
        row_data = [
            f.get("object", ""),
            f.get("field_api_name", ""),
            f.get("field_type", ""),
            f.get("rule_triggered", ""),
            status,
            f.get("description", ""),
            f.get("llm_suggested_description", ""),
            f.get("llm_reasoning", ""),
            "",   # Decision — Admin fills in
            "",   # Admin_Version — Admin fills in if Decision=Edit
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font, c.fill, c.border = CELL_FONT, fill, THIN_BORDER
            c.alignment = CENTER_AL if ci in (1, 3, 4, 5, 9) else LEFT_AL

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = _col_width(h)

    # Add dropdown validation for Decision column (col 9)
    try:
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(
            type="list",
            formula1='"Approve,Edit,Reject"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.sqref = f"I2:I{max(len(fields) + 1, 2)}"
        ws.add_data_validation(dv)
    except Exception:
        pass  # Non-critical: validation is cosmetic


def _write_reference_tab(wb, title: str, fields: list[dict]) -> None:
    ws = wb.create_sheet(title)
    headers = ["Object", "Field_API_Name", "Field_Type", "Original_Status",
               "Current_Description", "LLM_Action", "LLM_Reasoning"]
    _write_header_row(ws, headers)

    for ri, f in enumerate(fields, 2):
        status = f["classifier_status"]
        bg     = STATUS_COLORS.get(status, "FFFFFF")
        fill   = PatternFill("solid", fgColor=bg)
        row_data = [
            f.get("object", ""),
            f.get("field_api_name", ""),
            f.get("field_type", ""),
            status,
            f.get("description", ""),
            f.get("llm_action", "keep"),
            f.get("llm_reasoning", ""),
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font, c.fill, c.border = CELL_FONT, fill, THIN_BORDER
            c.alignment = CENTER_AL if ci in (1, 3, 4, 6) else LEFT_AL

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = _col_width(h)


def _write_summary_tab(wb, tab_a, tab_b, tab_c, timestamp: str) -> None:
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12

    rows = [
        ("Generated at (UTC)", timestamp),
        ("Tab A — FLAGGED", len(tab_a)),
        ("Tab B — UNCERTAIN / REVIEWED-rewrite", len(tab_b)),
        ("Tab C — Reference only", len(tab_c)),
        ("Total fields", len(tab_a) + len(tab_b) + len(tab_c)),
    ]
    for ri, (label, val) in enumerate(rows, 1):
        ws.cell(row=ri, column=1, value=label).font = Font(name="Arial", size=9, bold=True)
        ws.cell(row=ri, column=2, value=val).font   = Font(name="Arial", size=9)


def _write_header_row(ws, headers: list[str]) -> None:
    ws.row_dimensions[1].height = 28
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, CENTER_AL, THIN_BORDER


# ─────────────────────────────────────────────────────────────────────────────
# 12. UPDATE CACHE
# ─────────────────────────────────────────────────────────────────────────────

def update_cache(cache: dict, all_fields: list[dict], llm_results: dict) -> None:
    for f in all_fields:
        if f["classifier_status"] == "SKIPPED":
            continue
        key    = f"{f['object']}.{f['field_api_name']}"
        h      = description_hash(f.get("description", ""))
        result = llm_results.get(f["field_api_name"], {})
        cache[key] = {
            "hash":       h,
            "last_run":   datetime.utcnow().isoformat(),
            "status":     f["classifier_status"],
            "rule":       f.get("rule_triggered"),
            "llm_result": result,
        }
    save_cache(cache)


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    log.info("=" * 60)
    log.info("Script 1 — Ingest / Classify / Send")
    log.info("=" * 60)

    # 1. Config
    cfg = load_config()

    # 2. Ingest
    raw_fields = ingest_fields(cfg)

    # 3. Classify
    classified = classify_fields(raw_fields, cfg)

    # 4. Save sf_classified.json
    save_classified(classified)

    # 5. Cache
    cache    = load_cache()
    to_llm, already_cached = apply_cache(classified, cache)

    # 6. Prompts
    prompts = load_prompts()

    # 7. Sort + deduplicate
    unique_fields, dedup_name_map = sort_and_deduplicate(to_llm)

    # 8–9. LLM + save response
    all_llm_results: dict = {}
    if unique_fields:
        raw_results    = run_llm(unique_fields, prompts, cfg)
        all_llm_results = save_llm_response(raw_results, dedup_name_map)
    else:
        log.info("All fields served from cache — no LLM calls needed")
        save_llm_response({}, {})

    # Merge cached LLM results
    for f in already_cached:
        if f.get("llm_result"):
            all_llm_results[f["field_api_name"]] = f["llm_result"]

    # 10. Generate Excel
    generate_review_queue(classified, all_llm_results)

    # 11. Update cache
    update_cache(cache, classified, all_llm_results)

    log.info("=" * 60)
    log.info("Script 1 complete.")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
